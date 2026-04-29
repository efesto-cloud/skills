#!/usr/bin/env python3
"""
extract_map.py — Excel Mapper v2 extraction script
Extracts full structural data from an Excel file into a JSON file.

Usage:
    python extract_map.py <input.xlsx> <output.json> [--sample-only]
"""

import sys, os, re, json, subprocess
from collections import defaultdict
from datetime import datetime, date

FORMULA_SAMPLE_CAP = 100

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"],
        check=True
    )
    import openpyxl
    from openpyxl.utils import get_column_letter


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_str(v, max_len=40):
    if v is None: return ""
    s = str(v)
    return s[:max_len] + "…" if len(s) > max_len else s

def json_serializable(v):
    if isinstance(v, (datetime, date)): return str(v)
    return v

def detect_type(values):
    types = []
    for v in values:
        if v is None or v == "": continue
        if isinstance(v, bool): types.append("bool")
        elif isinstance(v, int): types.append("int")
        elif isinstance(v, float): types.append("float")
        elif isinstance(v, (datetime, date)): types.append("date")
        elif isinstance(v, str):
            s = v.strip().replace(",", ".")
            try: float(s); types.append("numeric-str")
            except ValueError: types.append("str")
        else: types.append("other")
    if not types: return "empty"
    counts = defaultdict(int)
    for t in types: counts[t] += 1
    if len(set(types)) > 2: return "mixed"
    return max(counts, key=counts.get)

def find_cross_sheet_refs(formula, sheet_names):
    if not formula: return set()
    refs = set()
    for name in sheet_names:
        # match SheetName! or 'Sheet Name'!
        pattern = r"(?:'?" + re.escape(name) + r"'?!)"
        if re.search(pattern, formula, re.IGNORECASE):
            refs.add(name)
    return refs

def detect_circular_refs(outgoing_refs: dict) -> list:
    """DFS cycle detection on the cross-sheet reference graph.
    Returns a list of cycle strings like ["A → B → A"]."""
    sys.setrecursionlimit(max(1000, len(outgoing_refs) * 10))
    visited, in_stack, path, cycles = set(), set(), [], []

    def dfs(node):
        if node in in_stack:
            cycle_start = path.index(node)
            cycles.append(" → ".join(path[cycle_start:] + [node]))
            return
        if node in visited:
            return
        visited.add(node)
        in_stack.add(node)
        path.append(node)
        for neighbor in outgoing_refs.get(node, set()):
            dfs(neighbor)
        path.pop()
        in_stack.discard(node)

    for sheet in list(outgoing_refs.keys()):
        if sheet not in visited:
            dfs(sheet)
    return cycles

def detect_external_links(formula):
    return bool(formula and "[" in formula and "]" in formula)

def count_formula_density(formulas_count, total_cells):
    if total_cells == 0: return 0.0
    return round(formulas_count / total_cells, 3)

def detect_layout(ws, min_row, max_row, min_col, max_col):
    """Heuristic: detect tabular vs block vs hybrid layout."""
    merged = list(ws.merged_cells.ranges)
    multi_row_merges = [m for m in merged if m.max_row > m.min_row]

    # Sample blank rows in data area
    blank_rows = 0
    sample_limit = min(max_row, min_row + 100)
    for r in range(min_row + 1, sample_limit + 1):
        row_vals = [ws.cell(r, c).value for c in range(min_col, min(min_col+5, max_col+1))]
        if all(v is None for v in row_vals):
            blank_rows += 1

    blank_ratio = blank_rows / max(sample_limit - min_row, 1)

    if multi_row_merges and blank_ratio > 0.05:
        return "block"
    elif multi_row_merges or blank_ratio > 0.03:
        return "hybrid"
    else:
        return "tabular"


# ── main extraction ───────────────────────────────────────────────────────────

def extract(input_path, output_path, sample_only=False):
    filename = os.path.basename(input_path)
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".csv":
        return extract_csv(input_path, output_path, filename)

    wb      = openpyxl.load_workbook(input_path, data_only=False)
    wb_data = openpyxl.load_workbook(input_path, data_only=True)
    sheet_names = wb.sheetnames

    # Named ranges
    named_ranges = []
    for name, obj in wb.defined_names.items():
        named_ranges.append({
            "name": name,
            "refers_to": obj.attr_text if hasattr(obj, "attr_text") else str(obj)
        })

    has_external_links = False
    total_formulas_global = 0
    sheets_data = []
    outgoing_refs  = defaultdict(set)   # sheet → {sheets it references}
    incoming_refs  = defaultdict(set)   # sheet → {sheets that reference it}

    for sheet_name in sheet_names:
        ws      = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        r1, r2 = ws.min_row, ws.max_row
        c1, c2 = ws.min_column, ws.max_column

        # Empty sheet guard
        if r1 is None or r2 is None or r1 == r2 == 1:
            cell_val = ws_data.cell(1, 1).value
            if cell_val is None:
                sheets_data.append({"name": sheet_name, "empty": True})
                continue

        row_limit = min(r2, r1 + 999) if sample_only else r2
        rows_truncated = sample_only and r2 > row_limit
        nrows = row_limit - r1 + 1
        ncols = c2 - c1 + 1
        total_cells = nrows * ncols

        layout = detect_layout(ws, r1, r2, c1, c2)

        # Headers from first row
        headers = {}
        for ci in range(c1, c2 + 1):
            v = ws.cell(r1, ci).value
            headers[ci] = safe_str(v) if v is not None else get_column_letter(ci)

        col_values   = defaultdict(list)
        formulas     = []
        formula_count = 0
        ext_link_flag = False

        for ri in range(r1 + 1, row_limit + 1):
            for ci in range(c1, c2 + 1):
                cell_f = ws.cell(ri, ci)
                cell_v = ws_data.cell(ri, ci)
                vf = cell_f.value
                vd = cell_v.value

                if isinstance(vf, str) and vf.startswith("="):
                    formula_count += 1
                    addr = f"{get_column_letter(ci)}{ri}"
                    if len(formulas) < FORMULA_SAMPLE_CAP:
                        formulas.append({"cell": addr, "formula": vf})
                    elif len(formulas) == FORMULA_SAMPLE_CAP:
                        print(f"  ⚠️  Sheet '{sheet_name}': formula sample capped at {FORMULA_SAMPLE_CAP} (more formulas exist)")
                    refs = find_cross_sheet_refs(vf, sheet_names)
                    outgoing_refs[sheet_name].update(refs - {sheet_name})
                    for r in refs - {sheet_name}:
                        incoming_refs[r].add(sheet_name)
                    if detect_external_links(vf):
                        ext_link_flag = True
                        has_external_links = True

                col_values[ci].append(vd)

        if ext_link_flag:
            has_external_links = True

        if rows_truncated:
            print(f"  ⚠️  Sheet '{sheet_name}': sampled {nrows} of {r2 - r1 + 1} rows (run without --sample-only for full extraction)")

        total_formulas_global += formula_count

        # Column summaries
        columns = []
        for ci in range(c1, c2 + 1):
            vals = col_values[ci]
            dtype = detect_type(vals)
            non_null = [v for v in vals if v is not None and v != ""]
            samples = []
            seen = set()
            for v in non_null:
                s = safe_str(v)
                if s not in seen:
                    seen.add(s); samples.append(s)
                if len(samples) >= 5: break

            columns.append({
                "letter": get_column_letter(ci),
                "header": headers.get(ci, ""),
                "dtype": dtype,
                "samples": samples
            })

        sheets_data.append({
            "name": sheet_name,
            "empty": False,
            "dimensions": f"{nrows} rows × {ncols} columns",
            "range": f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}",
            "layout": layout,
            "formula_count": formula_count,
            "formula_density": count_formula_density(formula_count, total_cells),
            "formulas": formulas,
            "columns": columns,
            "outgoing_refs": sorted(outgoing_refs[sheet_name]),
            "incoming_refs": sorted(incoming_refs[sheet_name]),
            "has_external_links": ext_link_flag,
            "sampled": sample_only
        })

    # Build final output
    circular_refs = detect_circular_refs(outgoing_refs)
    if circular_refs:
        print(f"  ⚠️  Circular references detected: {', '.join(circular_refs)}")

    data = {
        "filename": filename,
        "extracted_at": datetime.now().isoformat(),
        "sheet_count": len(sheet_names),
        "non_empty_sheets": sum(1 for s in sheets_data if not s.get("empty")),
        "has_external_links": has_external_links,
        "total_formulas": total_formulas_global,
        "named_ranges": named_ranges,
        "circular_refs": circular_refs,
        "has_circular_refs": bool(circular_refs),
        "sheets": sheets_data
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=json_serializable)

    print(f"✅ Extraction complete → {output_path}")
    print(f"   Sheets: {data['sheet_count']} ({data['non_empty_sheets']} non-empty)")
    print(f"   Formulas: {total_formulas_global}")
    print(f"   External links: {has_external_links}")
    print(f"   Circular refs: {circular_refs if circular_refs else 'none'}")
    return output_path


def extract_csv(input_path, output_path, filename):
    try:
        import pandas as pd
    except ImportError:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "pandas", "--break-system-packages", "-q"],
            check=True
        )
        import pandas as pd

    df = pd.read_csv(input_path, nrows=1000)
    columns = []
    for col in df.columns:
        non_null = df[col].dropna()
        samples = [str(v) for v in non_null.unique()[:5]]
        columns.append({
            "letter": "", "header": col,
            "dtype": str(df[col].dtype), "samples": samples
        })

    data = {
        "filename": filename,
        "extracted_at": datetime.now().isoformat(),
        "sheet_count": 1,
        "non_empty_sheets": 1,
        "has_external_links": False,
        "total_formulas": 0,
        "named_ranges": [],
        "sheets": [{
            "name": "CSV",
            "empty": False,
            "dimensions": f"{len(df)} rows × {len(df.columns)} columns",
            "range": f"A1:{get_column_letter(len(df.columns))}{len(df)+1}",
            "layout": "tabular",
            "formula_count": 0,
            "formula_density": 0.0,
            "formulas": [],
            "columns": columns,
            "outgoing_refs": [],
            "incoming_refs": [],
            "has_external_links": False,
            "sampled": True
        }]
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✅ CSV extraction complete → {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: extract_map.py <input> <output.json> [--sample-only]")
        sys.exit(1)
    extract(sys.argv[1], sys.argv[2], "--sample-only" in sys.argv)
